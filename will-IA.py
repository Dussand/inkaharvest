import streamlit as st
import pandas as pd
from supabase import create_client
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# SUPABASE_URL = "https://wkmulgceuhrvogzivyby.supabase.co"
# SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6IndrbXVsZ2NldWhydm9neml2eWJ5Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MDY4NjY0MiwiZXhwIjoyMDg2MjYyNjQyfQ.ITLUH4TNoScKTmTa7PSgyNqLYle6MTPKjo5bKsmRz_k"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

st.set_page_config(layout="wide")

st.title("📊 Facturas registradas para presentacion de impuestos")
st.subheader("🏭 WILLMACTEX S.A.C.")

# ── Estilos compartidos ──
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Red+Hat+Display:wght@400;600;700;800;900&display=swap');
.metric-card {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    border-radius: 18px;
    padding: 32px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    align-items: center;
    text-align: center;
    box-shadow: 0 8px 32px rgba(0,0,0,0.3), inset 0 1px 0 rgba(255,255,255,0.08);
    border: 1px solid rgba(255,255,255,0.08);
    font-family: 'Red Hat Display', sans-serif;
}
.metric-icon { font-size: 28px; margin-bottom: 12px; }
.metric-label {
    font-size: 12px;
    font-weight: 600;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: rgba(255,255,255,0.55);
    margin-bottom: 12px;
}
.metric-value {
    font-size: 38px;
    font-weight: 900;
    letter-spacing: -1px;
    color: #FFFFFF;
    line-height: 1.1;
}
.metric-card-sm {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    border-radius: 18px;
    padding: 28px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    align-items: center;
    text-align: center;
    box-shadow: 0 8px 32px rgba(0,0,0,0.3), inset 0 1px 0 rgba(255,255,255,0.08);
    border: 1px solid rgba(255,255,255,0.08);
    font-family: 'Red Hat Display', sans-serif;
    transition: 0.2s ease;
}
.metric-card-sm:hover { transform: translateY(-3px); }
.metric-label-sm {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: rgba(255,255,255,0.55);
    margin-bottom: 8px;
}
.metric-value-sm {
    font-size: 30px;
    font-weight: 900;
    letter-spacing: -1px;
    color: #FFFFFF;
}
</style>
""", unsafe_allow_html=True)

def fetch_comprobantes():
    response = (
        supabase
        .table("comprobantes")
        .select("*")
        .order("created_at", desc=False)
        .execute()
    )
    return pd.DataFrame(response.data)

try:
    df_raw = fetch_comprobantes()

    # ── Selector de periodo ──
    hoy = pd.Timestamp.today()
    mes_presentacion = hoy - pd.DateOffset(months=1)
    año_base = mes_presentacion.year

    periodos = [f'{año_base}{str(m).zfill(2)}' for m in range(1, 13)]
    index_default = periodos.index(mes_presentacion.strftime('%Y%m'))

    periodo_seleccionado = st.selectbox('Seleccionar periodo', periodos, index=index_default)

    # ── Preparar datos base ──
    df_raw['fecha_emision'] = pd.to_datetime(df_raw['fecha_emision'], errors='coerce', dayfirst=True)
    df_raw['periodo'] = df_raw['fecha_emision'].dt.strftime("%Y%m")
    df_filtrado = df_raw[df_raw['periodo'] == periodo_seleccionado].copy()

    # Columnas a mostrar (orden definido)
    COLS_MOSTRAR = ['fecha_emision', 'empresa', 'serie', 'nfactura', 'ruc_comprador', 'base_imp', 'igv', 'total']

    #########################################
    ## VENTAS
    #########################################
    st.divider()
    st.subheader("📤 Ventas - WILLMACTEX S.A.C.")

    ventas = df_filtrado[df_filtrado['tipo'] == 'venta'].copy()
    cols_ventas = [c for c in COLS_MOSTRAR if c in ventas.columns]
    ventas_show = ventas[cols_ventas]

    suma_base_ventas  = round(ventas['base_imp'].fillna(0).sum(), 2)
    suma_igv_ventas   = round(ventas['igv'].fillna(0).sum(), 2)
    suma_total_ventas = round(ventas['total'].fillna(0).sum(), 2)
    renta_a_pagar     = round(suma_base_ventas * 0.015, 2)
    cantidad_ventas   = len(ventas)

    altura_ventas = min(38 + len(ventas_show) * 36, 600)

    col_tv, col_mv = st.columns([3, 2], gap="large")

    with col_tv:
        st.dataframe(ventas_show, use_container_width=True, height=altura_ventas)

    with col_mv:
        st.markdown(f"""
        <div style="display:flex; flex-direction:column; gap:14px;">
            <div class="metric-card" style="padding:28px 20px;">
                <div class="metric-icon">💰</div>
                <div class="metric-label">Venta Total (Base Imp.)</div>
                <div class="metric-value">S/ {suma_base_ventas:,.2f}</div>
            </div>
            <div class="metric-card" style="padding:28px 20px;">
                <div class="metric-icon">🧾</div>
                <div class="metric-label">Renta a Pagar (1.5%)</div>
                <div class="metric-value">S/ {renta_a_pagar:,.2f}</div>
            </div>
            <div class="metric-card" style="padding:28px 20px;">
                <div class="metric-icon">📊</div>
                <div class="metric-label">Total Comprobantes</div>
                <div class="metric-value">{cantidad_ventas}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    #########################################
    ## COMPRAS
    #########################################
    st.divider()
    st.subheader("📥 Compras - WILLMACTEX S.A.C.")

    compras = df_filtrado[df_filtrado['tipo'] == 'compra'].copy()

    # Para compras mostramos ruc_vendedor en lugar de ruc_comprador
    COLS_COMPRAS = ['fecha_emision', 'empresa', 'serie', 'nfactura', 'ruc_vendedor', 'base_imp', 'igv', 'total']
    cols_compras = [c for c in COLS_COMPRAS if c in compras.columns]
    compras_show = compras[cols_compras]

    suma_base_compras  = round(compras['base_imp'].fillna(0).sum(), 2)
    suma_igv_compras   = round(compras['igv'].fillna(0).sum(), 2)
    suma_total_compras = round(compras['total'].fillna(0).sum(), 2)
    cantidad_compras   = len(compras)

    altura_compras = max(min(38 + len(compras_show) * 36, 600), 250)

    col_tc, col_mc = st.columns([3, 2], gap="large")

    with col_tc:
        st.dataframe(compras_show, use_container_width=True, height=altura_compras)

    with col_mc:
        st.markdown(f"""
        <div style="display:flex; flex-direction:column; gap:14px;">
            <div style="display:flex; gap:14px;">
                <div class="metric-card-sm" style="flex:1; padding:24px 16px;">
                    <div class="metric-label-sm">💵 Base Imponible</div>
                    <div class="metric-value-sm">S/ {suma_base_compras:,.2f}</div>
                </div>
                <div class="metric-card-sm" style="flex:1; padding:24px 16px;">
                    <div class="metric-label-sm">🧾 IGV Compras</div>
                    <div class="metric-value-sm">S/ {suma_igv_compras:,.2f}</div>
                </div>
            </div>
            <div style="display:flex; gap:14px;">
                <div class="metric-card-sm" style="flex:1; padding:24px 16px;">
                    <div class="metric-label-sm">🏷️ Total Compras</div>
                    <div class="metric-value-sm">S/ {suma_total_compras:,.2f}</div>
                </div>
                <div class="metric-card-sm" style="flex:1; padding:24px 16px;">
                    <div class="metric-label-sm">📦 N° Facturas</div>
                    <div class="metric-value-sm">{cantidad_compras}</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    #########################################
    ## EXPORTAR EXCEL
    #########################################
    st.divider()

    def generar_excel(ventas_df, compras_df, igv_ventas, igv_compras, renta):
        wb = Workbook()

        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', start_color='0F3460')
        center      = Alignment(horizontal='center', vertical='center')
        money_fmt   = '#,##0.00'
        thin        = Side(style='thin', color='CCCCCC')
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        def escribir_hoja(ws, df, titulo):
            ws.append([titulo])
            ws['A1'].font = Font(name='Arial', bold=True, size=13)
            ws.append([])

            headers = list(df.columns)
            ws.append(headers)
            for col_idx, _ in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center
                cell.border    = border

            for row_data in df.itertuples(index=False):
                ws.append(list(row_data))

            for col_idx, col_name in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 22
                if df[col_name].dtype in ['float64', 'int64']:
                    for row_idx in range(4, 4 + len(df)):
                        ws.cell(row=row_idx, column=col_idx).number_format = money_fmt
                for row_idx in range(4, 4 + len(df)):
                    ws.cell(row=row_idx, column=col_idx).border = border

        # Hoja 1: Ventas
        ws1 = wb.active
        ws1.title = 'Ventas'
        escribir_hoja(ws1, ventas_df, 'Reporte de Ventas - WILLMACTEX S.A.C.')

        # Hoja 2: Compras
        ws2 = wb.create_sheet('Compras')
        escribir_hoja(ws2, compras_df, 'Reporte de Compras - WILLMACTEX S.A.C.')

        # Hoja 3: Liquidación
        ws3 = wb.create_sheet('Liquidacion')
        ws3['A1'] = 'Liquidación de Impuestos - WILLMACTEX S.A.C.'
        ws3['A1'].font = Font(name='Arial', bold=True, size=13)

        ws3.append([])
        ws3.append(['Concepto', 'Monto (S/)'])
        for col_idx in range(1, 3):
            cell = ws3.cell(row=3, column=col_idx)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border

        igv_pagar = round(igv_ventas - igv_compras, 2)
        filas_liq = [
            ['IGV Ventas',    igv_ventas],
            ['IGV Compras',   igv_compras],
            ['IGV a Pagar',   igv_pagar],
            ['Renta a Pagar (1.5%)', renta],
        ]
        for i, fila in enumerate(filas_liq, start=4):
            ws3.cell(row=i, column=1, value=fila[0]).border = border
            cell_val = ws3.cell(row=i, column=2, value=fila[1])
            cell_val.number_format = money_fmt
            cell_val.border = border
            if fila[0] in ('IGV a Pagar', 'Renta a Pagar (1.5%)'):
                cell_val.font = Font(name='Arial', bold=True, color='0F3460')

        ws3.column_dimensions['A'].width = 28
        ws3.column_dimensions['B'].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    excel_buffer = generar_excel(
        ventas_show,
        compras_show,
        suma_igv_ventas,
        suma_igv_compras,
        renta_a_pagar
    )

    st.download_button(
        label="📥 Exportar Liquidación Excel",
        data=excel_buffer,
        file_name=f"liquidacion-WILLMACTEX-{periodo_seleccionado}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

except Exception as e:
    st.error(f"Error cargando datos: {e}")
import streamlit as st
import pandas as pd
from supabase import create_client
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

st.set_page_config(layout="wide")

st.title("📊 Facturas registradas para presentacion de impuestos")

def fetch_table(table_name):
    response = (
        supabase
        .table(table_name)
        .select("*")
        .order("created_at", desc=False)
        .execute()
    )
    return pd.DataFrame(response.data)

try:
    compras_inka = fetch_table("compras_inkahavrvest")
    ventas_inka = fetch_table("comprobantes_inkah")

    hoy = pd.Timestamp.today()
    mes_presentacion = hoy - pd.DateOffset(months=1)
    año_base = mes_presentacion.year
    periodo_actual = mes_presentacion.strftime('%Y%m')

    periodos = [f'{año_base}{str(m).zfill(2)}' for m in range(1, 13)]
    index_default = periodos.index(mes_presentacion.strftime('%Y%m'))

    periodo_seleccionado = st.selectbox('Seleccionar periodo', periodos, index=index_default)
#########################################
    ## SHOW REPORTS SELLS
    #########################################

    st.subheader("📤 Ventas - Inkaharvest")

    columns_drop = {'id', 'ruc_comprador', 'tipo_documento', 'empresa'}
    ventas_inka.drop(columns=columns_drop, inplace=True)

    ventas_inka['fecha_emision'] = pd.to_datetime(ventas_inka['fecha_emision'], errors='coerce')
    ventas_inka['periodo'] = ventas_inka['fecha_emision'].dt.strftime("%Y%m")

    cols_hide_ventas = ['fecha_inicio', 'fecha_fin', 'created_at']
    cols_to_show_ventas = [c for c in ventas_inka.columns if c not in cols_hide_ventas]
    ventas_inka = ventas_inka[cols_to_show_ventas]
    ventas_inka = ventas_inka[ventas_inka['periodo'] == periodo_seleccionado]

    suma_baseImponible = ventas_inka['base_imponible'].fillna(0).sum()
    renta_aPagar = round(suma_baseImponible * 0.01, 0)
    cantidad_facturas = len(ventas_inka)
    suma_igv = ventas_inka['igv'].fillna(0).sum()

    # Altura dinámica basada en filas
    altura_ventas = min(38 + len(ventas_inka) * 36, 600)

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Red+Hat+Display:wght@400;600;700;800;900&display=swap');
    .metric-card {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        border-radius: 18px;
        padding: 32px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
        text-align: center;
        box-shadow: 0 8px 32px rgba(0,0,0,0.3), inset 0 1px 0 rgba(255,255,255,0.08);
        border: 1px solid rgba(255,255,255,0.08);
        font-family: 'Red Hat Display', sans-serif;
    }
    .metric-icon { font-size: 28px; margin-bottom: 12px; }
    .metric-label {
        font-size: 12px;
        font-weight: 600;
        letter-spacing: 2px;
        text-transform: uppercase;
        color: rgba(255,255,255,0.55);
        margin-bottom: 12px;
    }
    .metric-value {
        font-size: 38px;
        font-weight: 900;
        letter-spacing: -1px;
        color: #FFFFFF;
        line-height: 1.1;
    }
    .metric-card-sm {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        border-radius: 18px;
        padding: 28px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
        text-align: center;
        box-shadow: 0 8px 32px rgba(0,0,0,0.3), inset 0 1px 0 rgba(255,255,255,0.08);
        border: 1px solid rgba(255,255,255,0.08);
        font-family: 'Red Hat Display', sans-serif;
        transition: 0.2s ease;
    }
    .metric-card-sm:hover { transform: translateY(-3px); }
    .metric-label-sm {
        font-size: 11px;
        font-weight: 600;
        letter-spacing: 2px;
        text-transform: uppercase;
        color: rgba(255,255,255,0.55);
        margin-bottom: 8px;
    }
    .metric-value-sm {
        font-size: 30px;
        font-weight: 900;
        letter-spacing: -1px;
        color: #FFFFFF;
    }
    </style>
    """, unsafe_allow_html=True)

    col_tabla_v, col_metrics_v = st.columns([3, 2], gap="large")

    with col_tabla_v:
        st.dataframe(ventas_inka, use_container_width=True, height=altura_ventas)

    with col_metrics_v:
        st.markdown(f"""
        <div style="display:flex; flex-direction:column; gap:12px; height:{altura_ventas}px;">
            <div class="metric-card" style="flex:1;">
                <div class="metric-icon">💰</div>
                <div class="metric-label">Venta Total</div>
                <div class="metric-value">{suma_baseImponible:,.2f}</div>
            </div>
            <div class="metric-card" style="flex:1;">
                <div class="metric-icon">🧾</div>
                <div class="metric-label">Impuesto a Pagar</div>
                <div class="metric-value">{renta_aPagar:,.2f}</div>
            </div>
            <div class="metric-card" style="flex:1;">
                <div class="metric-icon">📊</div>
                <div class="metric-label">Total de Comprobantes</div>
                <div class="metric-value">{cantidad_facturas}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    #########################################
    ## SHOW REPORTS PURCHASES
    #########################################
    st.divider()
    st.subheader("📥 Compras - Inkaharvest")
    st.write('''Facturas listas para presentar que incluyen las facturas sin detraccion y 
    las facturas que tienen detraccion y que ya han sido pagadas dentro del periodo
    establecido por sunat que es hasta el 5to dia habil del mes siguiente.''')

    compras_inka['fecha_emision'] = pd.to_datetime(compras_inka['fecha_emision'], errors='coerce')
    compras_inka['periodo'] = compras_inka['fecha_emision'].dt.strftime("%Y%m")

    cols_hide_compras = ['id', 'fecha_inicio', 'fecha_fin', 'created_at', 'detalle_compra']
    cols_to_show_compras = [c for c in compras_inka.columns if c not in cols_hide_compras]
    compras_inka = compras_inka[cols_to_show_compras]
    compras_inka = compras_inka[compras_inka['periodo'] == periodo_seleccionado]

    compras_inka_presentar = compras_inka[
        ((compras_inka['detraccion'] == 'SI') & (compras_inka['estado'] == 'PAGADO')) |
        (compras_inka['detraccion'] == 'NO')
    ]

    total_bigrav    = round(compras_inka['BI_gravado'].fillna(0).sum(), 2)
    total_bi_nograv = round(compras_inka['BI_nogravado'].fillna(0).sum(), 2)
    total_igvgrav   = round(compras_inka['igv_gravado'].fillna(0).sum(), 2)
    total_igv_no    = round(compras_inka['igv_nogravado'].fillna(0).sum(), 2)
    total_grav      = round(compras_inka['total_gravado'].fillna(0).sum(), 2)
    total_nograv    = round(compras_inka['total_nogravado'].fillna(0).sum(), 2)

# Altura dinámica basada en filas
    altura_compras = max(min(38 + len(compras_inka_presentar) * 36, 600), 350)

    col_tabla_c, col_metrics_c = st.columns([3, 2], gap="large")

    with col_tabla_c:
        st.dataframe(compras_inka_presentar, use_container_width=True, height=altura_compras)

    with col_metrics_c:
        st.markdown(f"""
        <div style="display:flex; flex-direction:column; gap:12px; height:{altura_compras}px; overflow:hidden;">
            <div style="display:flex; gap:12px; flex:1; min-height:0;">
                <div class="metric-card-sm" style="flex:1;">
                    <div class="metric-label-sm">💵 BI Gravado</div>
                    <div class="metric-value-sm">{total_bigrav:,.2f}</div>
                </div>
                <div class="metric-card-sm" style="flex:1;">
                    <div class="metric-label-sm">📦 BI No Gravado</div>
                    <div class="metric-value-sm">{total_bi_nograv:,.2f}</div>
                </div>
            </div>
            <div style="display:flex; gap:12px; flex:1; min-height:0;">
                <div class="metric-card-sm" style="flex:1;">
                    <div class="metric-label-sm">🧾 IGV Gravado</div>
                    <div class="metric-value-sm">{total_igvgrav:,.2f}</div>
                </div>
                <div class="metric-card-sm" style="flex:1;">
                    <div class="metric-label-sm">💰 Total Gravado</div>
                    <div class="metric-value-sm">{total_grav:,.2f}</div>
                </div>
            </div>
            <div style="display:flex; gap:12px; flex:1; min-height:0;">
                <div class="metric-card-sm" style="flex:1;">
                    <div class="metric-label-sm">📊 Total No Gravado</div>
                    <div class="metric-value-sm">{total_nograv:,.2f}</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.divider()
    st.subheader("📥 Compras a revisar pago de detraccion - Inkaharvest")

    compras_inka_revisar = compras_inka[
        ((compras_inka['detraccion'] == 'SI') & (compras_inka['estado'] == 'PENDIENTE'))
    ]
    if compras_inka_revisar.empty:
        st.success("✅ No hay compras pendientes de revisión")
    else:
        compras_inka_revisar.insert(0, 'pagado', False)

        editada = st.data_editor(
            compras_inka_revisar,
            use_container_width=True,
            column_config={
                "pagado": st.column_config.CheckboxColumn("¿Pagado?", default=False)
            },
            disabled=[c for c in compras_inka_revisar.columns if c != 'pagado'],
            hide_index=True
        )

        filas_marcadas = editada[editada['pagado'] == True]

        if not filas_marcadas.empty:
            if st.button(f"✅ Confirmar pago de {len(filas_marcadas)} detraccion(es)", use_container_width=True):
                for _, row in filas_marcadas.iterrows():
                    supabase.table("compras_inkahavrvest") \
                        .update({"estado": "PAGADO"}) \
                        .eq("numero", row["numero"]) \
                        .execute()
                st.success("✅ Detracciones actualizadas correctamente.")
                st.rerun()
    #########################################
    ## EXPORTAR EXCEL
    #########################################
    st.divider()

    def generar_excel(ventas_df, compras_df, suma_igv, igv_compras, renta):
        wb = Workbook()

        # ── Estilos ──
        header_font     = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        header_fill     = PatternFill('solid', start_color='0F3460')
        center          = Alignment(horizontal='center', vertical='center')
        money_fmt       = '#,##0.00'
        thin            = Side(style='thin', color='CCCCCC')
        border          = Border(left=thin, right=thin, top=thin, bottom=thin)

        def escribir_hoja(ws, df, titulo):
            ws.append([titulo])
            ws['A1'].font = Font(name='Arial', bold=True, size=13)
            ws.append([])  # fila vacía

            # Header
            headers = list(df.columns)
            ws.append(headers)
            for col_idx, _ in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.font     = header_font
                cell.fill     = header_fill
                cell.alignment = center
                cell.border   = border

            # Datos
            for row_data in df.itertuples(index=False):
                ws.append(list(row_data))

            # Formato columnas numéricas + ancho
            for col_idx, col_name in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 20
                if df[col_name].dtype in ['float64', 'int64']:
                    for row_idx in range(4, 4 + len(df)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.number_format = money_fmt
                for row_idx in range(4, 4 + len(df)):
                    ws.cell(row=row_idx, column=col_idx).border = border

        # ── Hoja 1: Ventas ──
        ws1 = wb.active
        ws1.title = 'Ventas'
        escribir_hoja(ws1, ventas_df, 'Reporte de Ventas')

        # ── Hoja 2: Compras ──
        ws2 = wb.create_sheet('Compras')
        escribir_hoja(ws2, compras_df, 'Reporte de Compras')

        # ── Hoja 3: Liquidación ──
        ws3 = wb.create_sheet('Liquidacion')
        ws3['A1'] = 'Liquidación de Impuestos'
        ws3['A1'].font = Font(name='Arial', bold=True, size=13)

        headers_liq = ['Concepto', 'Monto (S/)']
        ws3.append([])
        ws3.append(headers_liq)
        for col_idx in range(1, 3):
            cell = ws3.cell(row=3, column=col_idx)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border

        igv_pagar = round(suma_igv - igv_compras, 2)
        filas_liq = [
            ['IGV Ventas',      suma_igv],
            ['IGV Compras',     igv_compras],
            ['IGV a Pagar',     igv_pagar],
            ['Renta a Pagar',   renta],
        ]
        for i, fila in enumerate(filas_liq, start=4):
            ws3.cell(row=i, column=1, value=fila[0]).border = border
            cell_val = ws3.cell(row=i, column=2, value=fila[1])
            cell_val.number_format = money_fmt
            cell_val.border = border
            if fila[0] == 'IGV a Pagar':
                cell_val.font = Font(name='Arial', bold=True, color='0F3460')

        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 20

        # ── Guardar en buffer ──
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    excel_buffer = generar_excel(
        ventas_inka,
        compras_inka_presentar,
        suma_igv,   # IGV ventas — ajusta si tienes variable propia
        total_igvgrav,
        renta_aPagar
    )

    st.download_button(
        label="📥 Exportar Liquidación Excel",
        data=excel_buffer,
        file_name=f"liquidacionImp-{periodo_seleccionado}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
except Exception as e:
    st.error(f"Error cargando datos: {e}")